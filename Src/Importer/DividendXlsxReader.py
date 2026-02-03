from __future__ import annotations
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook

from DividendReader import DividendReader
from Model.DividendRow import DividendRow


class DividendXlsxReader(DividendReader):
    def read(self, input_path: str) -> List[DividendRow]:
        workbook = load_workbook(filename=input_path, data_only=True)
        worksheet = workbook.active

        header_map = self._read_header_map(worksheet)
        if header_map is None:
            return []

        rows: List[DividendRow] = []

        for row_index in range(2, worksheet.max_row + 1):
            raw_row = self._read_row_as_dict(worksheet, row_index, header_map)
            if raw_row is None:
                continue

            dividend_row = self._parse_row(raw_row)
            if dividend_row is None:
                continue

            rows.append(dividend_row)

        return rows

    def _read_header_map(self, worksheet: Any) -> Optional[Dict[str, int]]:
        """
        Returns a mapping of header name -> column index (1-based).
        """
        header_map: Dict[str, int] = {}

        for col_index in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col_index).value
            if cell_value is None:
                continue

            header = str(cell_value).strip()
            if not header:
                continue

            header_map[header] = col_index

        if not header_map:
            return None

        return header_map

    def _read_row_as_dict(
        self,
        worksheet: Any,
        row_index: int,
        header_map: Dict[str, int]
    ) -> Optional[Dict[str, Any]]:
        """
        Reads a worksheet row into a dict using the header row names.
        If the row is completely empty, returns None.
        """
        raw_row: Dict[str, Any] = {}
        any_value = False

        for header, col_index in header_map.items():
            value = worksheet.cell(row=row_index, column=col_index).value
            raw_row[header] = value

            if value is not None and str(value).strip() != "":
                any_value = True

        if not any_value:
            return None

        return raw_row

    def _parse_row(self, raw_row: Dict[str, Any]) -> Optional[DividendRow]:
        ticker = self._parse_ticker(raw_row)
        if ticker is None:
            return None

        trans_date = self._parse_date(raw_row)
        if trans_date is None:
            return None

        amount = self._parse_amount(raw_row)
        if amount is None:
            return None

        description = self._get_string(raw_row, ["Description", "Action"]) or "Dividend"
        currency = self._get_string(raw_row, ["Currency"]) or "CAD"

        return DividendRow(
            ticker=ticker,
            trans_date=trans_date,
            amount=amount,
            description=description.strip(),
            currency=currency.strip().upper(),
        )

    def _parse_ticker(self, raw_row: Dict[str, Any]) -> Optional[str]:
        ticker_value = self._get_string(raw_row, ["Symbol", "Ticker", "symbol"])
        if ticker_value is None:
            return None

        ticker = ticker_value.strip().upper()
        if not ticker:
            return None

        return ticker

    def _parse_date(self, raw_row: Dict[str, Any]) -> Optional[date]:
        value = self._get_value(raw_row, ["Transaction Date", "Settlement Date", "Date"])
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        date_str = str(value).strip()
        if not date_str:
            return None

        date_formats = [
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%Y-%m-%d %I:%M:%S %p",
        ]

        for date_format in date_formats:
            try:
                return datetime.strptime(date_str, date_format).date()
            except ValueError:
                continue

        return None

    def _parse_amount(self, raw_row: Dict[str, Any]) -> Optional[Decimal]:
        value = self._get_value(raw_row, ["Net Amount", "Amount", "amount"])
        if value is None:
            return None

        if isinstance(value, (int, float, Decimal)):
            try:
                return Decimal(str(value))
            except Exception:
                return None

        amount_str = str(value).replace("$", "").replace(",", "").strip()
        if not amount_str:
            return None

        try:
            return Decimal(amount_str)
        except Exception:
            return None

    def _get_string(self, raw_row: Dict[str, Any], keys: List[str]) -> Optional[str]:
        value = self._get_value(raw_row, keys)
        if value is None:
            return None

        return str(value)

    def _get_value(self, raw_row: Dict[str, Any], keys: List[str]) -> Optional[Any]:
        for key in keys:
            if key in raw_row and raw_row[key] is not None:
                return raw_row[key]

        return None
